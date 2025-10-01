# produccion/views.py

import decimal
import json
import io
import base64
from datetime import datetime, timedelta
from django.db.models import Sum
from openpyxl import Workbook
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Producto, Insumo, Formulacion, PasoDeProduccion, Proceso, Venta, VentaItem
from .forms import ProductoForm, FormulacionForm, InsumoForm, FormulacionUpdateForm, ProcesoForm, PasoUpdateForm, PasoDeProduccionForm, CalculadoraLotesForm, VentaItemFormSet
from cuentas.decorators import rol_requerido, mipyme_requerida
from cuentas.forms import CambiarContrasenaForm, ActualizarPerfilForm, ConfigurarAvatarForm, EditarInformacionEmpresaForm, ConfigurarLogoForm, CambiarSectorEconomicoForm, ConfigurarParametrosProduccionForm
from cuentas.models import Usuario
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages


@login_required
@mipyme_requerida
def panel_produccion(request):
    """
    Vista principal de la aplicación 'produccion'.
    Solo accesible para usuarios con una Mipyme asociada.
    """
    # Gracias al decorador, ahora podemos estar SEGUROS de que
    # request.user.mipyme existe y no es None.

    mipyme = request.user.mipyme

    # Calcular costos mensuales para umbrales
    num_usuarios = Usuario.objects.filter(mipyme=mipyme).count()
    costo_admin = decimal.Decimal(365)  # Salario admin
    costo_empleados = decimal.Decimal(350) * (num_usuarios - 1) if num_usuarios > 1 else decimal.Decimal(0)
    costos_mensuales = costo_admin + costo_empleados
    umbral_perdidas = -float(costos_mensuales)  # Línea roja para pérdidas
    umbral_ganancias = float(costos_mensuales)  # Línea verde para ganancias

    # Obtener ventas agrupadas por mes, últimos 12 meses con datos
    from django.db.models.functions import TruncMonth
    ventas_por_mes = Venta.objects.filter(mipyme=mipyme).annotate(
        mes=TruncMonth('fecha')
    ).values('mes').annotate(
        total_ventas=Sum('total')
    ).order_by('-mes')[:12]

    meses = []
    rentabilidades = []

    for venta_mes in reversed(list(ventas_por_mes)):
        mes = venta_mes['mes']
        ventas_mes = venta_mes['total_ventas'] or decimal.Decimal(0)
        rentabilidad = float(ventas_mes - costos_mensuales)

        meses.append(mes.strftime('%b %Y'))
        rentabilidades.append(rentabilidad)

    # Si no hay ventas, mostrar mensaje o gráfica vacía
    if not meses:
        meses = ['Sin datos']
        rentabilidades = [0]

    # Preparar datos para tabla alternativa
    datos_tabla = []
    for i, mes in enumerate(meses):
        datos_tabla.append({
            'mes': mes,
            'rentabilidad': rentabilidades[i] if i < len(rentabilidades) else 0,
            'estado': 'ganancia' if rentabilidades[i] > umbral_ganancias else ('perdida' if rentabilidades[i] < umbral_perdidas else 'neutral')
        })
    
    # Intentar generar gráfica con matplotlib
    grafica_generada = False
    error_grafica = None
    
    try:
        import matplotlib
        matplotlib.use('Agg')  # Usar backend no interactivo
        import matplotlib.pyplot as plt
        import numpy as np

        # Generar gráfica
        rentabilidades_array = np.array(rentabilidades)
        plt.figure(figsize=(10, 5))
        plt.plot(meses, rentabilidades, marker='o', color='blue', linewidth=2, label='Rentabilidad')
        plt.axhline(y=umbral_perdidas, color='red', linestyle='--', linewidth=2, label=f'Umbral Pérdidas (${umbral_perdidas:.0f})')
        plt.axhline(y=umbral_ganancias, color='green', linestyle='--', linewidth=2, label=f'Umbral Ganancias (${umbral_ganancias:.0f})')
        plt.fill_between(meses, umbral_perdidas, rentabilidades_array, where=(rentabilidades_array < umbral_perdidas), color='red', alpha=0.3)
        plt.fill_between(meses, umbral_ganancias, rentabilidades_array, where=(rentabilidades_array > umbral_ganancias), color='green', alpha=0.3)
        plt.title('Rentabilidad del Negocio - Últimos 12 Meses')
        plt.xlabel('Mes')
        plt.ylabel('Rentabilidad ($)')
        plt.xticks(rotation=45)
        plt.grid(True)
        plt.legend()
        plt.tight_layout()

        # Convertir a base64
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        image_base64 = base64.b64encode(buf.read()).decode('utf-8')
        buf.close()
        plt.close()
        
        grafica_generada = True
        grafica_rentabilidad = image_base64
        
    except ImportError as e:
        # Si matplotlib no está disponible o hay problemas con librerías del sistema
        error_msg = str(e)
        if 'libstdc++' in error_msg:
            error_grafica = "No se puede generar la gráfica debido a librerías del sistema faltantes (libstdc++). Se muestra tabla de datos como alternativa."
        else:
            error_grafica = f"No se pudo cargar matplotlib: {error_msg}"
        grafica_rentabilidad = None
        
    except Exception as e:
        # Capturar cualquier otro error durante la generación de la gráfica
        error_grafica = f"Error al generar la gráfica: {str(e)}"
        grafica_rentabilidad = None
    
    # Preparar datos para Chart.js
    chart_data = {
        'labels': meses,
        'datasets': [{
            'label': 'Rentabilidad ($)',
            'data': rentabilidades,
            'borderColor': 'rgba(75, 192, 192, 1)',
            'backgroundColor': 'rgba(75, 192, 192, 0.2)',
            'borderWidth': 2,
            'pointBackgroundColor': 'rgba(75, 192, 192, 1)',
            'pointRadius': 4,
            'tension': 0.1
        }]
    }

    # Construir contexto
    contexto = {
        'titulo': 'Panel de Producción',
        'usuario': request.user,
        'nombrepine': mipyme.nombre,
        'chart_data': json.dumps(chart_data),
        'error_grafica': error_grafica,
        'datos_tabla': datos_tabla,
        'umbral_perdidas': umbral_perdidas,
        'umbral_ganancias': umbral_ganancias,
        'tiene_datos': len(datos_tabla) > 0 and datos_tabla[0]['mes'] != 'Sin datos'
    }

    return render(request, 'produccion/panel.html', contexto)
@login_required
def lista_productos(request):
    """
    Muestra una lista de todos los productos pertenecientes a la Mipyme del usuario logueado.
    """
    # Asumimos que el modelo Mipyme está relacionado con el User (ej. OneToOneField)
    # y se puede acceder a través de request.user.mipyme
    try:
        mipyme_usuario = request.user.mipyme
        productos = Producto.objects.filter(mipyme=mipyme_usuario).order_by('nombre')
    except AttributeError:
        # Manejar el caso de que el usuario no tenga una Mipyme asociada
        # (esto no debería pasar si la lógica de cuentas es correcta)
        productos = []

    contexto = {
        'productos': productos,
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/lista_productos.html', contexto)


@login_required
@mipyme_requerida
@rol_requerido('ADMIN', 'EDITOR')
def crear_producto(request):
    """
    Gestiona la creación de un nuevo producto.
    """
    mipyme = request.user.mipyme
    usar_porcentaje_predeterminado = mipyme.porcentaje_ganancia_predeterminado > 0

    if request.method == 'POST':
        # Si el formulario fue enviado, procesa los datos
        form = ProductoForm(request.POST, usar_porcentaje_predeterminado=usar_porcentaje_predeterminado, porcentaje_predeterminado=mipyme.porcentaje_ganancia_predeterminado)
        if form.is_valid():
            # El formulario es válido, pero no lo guardes todavía
            nuevo_producto = form.save(commit=False)

            # Asigna la Mipyme del usuario actual al nuevo producto
            # Esto es crucial para la seguridad y la lógica de negocio
            nuevo_producto.mipyme = mipyme

            # Si se usa porcentaje predeterminado, asignarlo automáticamente
            if usar_porcentaje_predeterminado:
                nuevo_producto.porcentaje_ganancia = mipyme.porcentaje_ganancia_predeterminado

            # Ahora sí, guarda el objeto completo en la base de datos
            nuevo_producto.save()

            # Redirige al usuario a la lista de productos para que vea el nuevo item
            return redirect('produccion:lista_productos')
    else:
        # Si la petición es GET, muestra un formulario vacío
        form = ProductoForm(usar_porcentaje_predeterminado=usar_porcentaje_predeterminado, porcentaje_predeterminado=mipyme.porcentaje_ganancia_predeterminado)

    contexto = {
        'form': form,
        'usar_porcentaje_predeterminado': usar_porcentaje_predeterminado,
        'porcentaje_predeterminado': mipyme.porcentaje_ganancia_predeterminado,
        'nombrepine': mipyme.nombre
    }
    return render(request, 'produccion/crear_producto.html', contexto)


@login_required
@mipyme_requerida
def detalle_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id, mipyme=request.user.mipyme)
    mipyme_actual = request.user.mipyme

    # Determinar si usar el margen de desperdicio predeterminado
    usar_margen_desperdicio_predeterminado = mipyme_actual.margen_desperdicio_predeterminado > 0

    # Inicializamos ambos formularios para pasarlos al contexto
    form_insumo = FormulacionForm(
        mipyme=mipyme_actual,
        usar_margen_desperdicio_predeterminado=usar_margen_desperdicio_predeterminado
    )
    form_proceso = PasoDeProduccionForm(mipyme=mipyme_actual)

    if request.method == 'POST':
        # Identificamos qué formulario se envió usando el nombre del botón 'submit'

        # --- LÓGICA PARA EL FORMULARIO DE INSUMOS ---
        if 'submit_insumo' in request.POST:
            form_insumo = FormulacionForm(
                request.POST,
                mipyme=mipyme_actual,
                usar_margen_desperdicio_predeterminado=usar_margen_desperdicio_predeterminado
            )
            if form_insumo.is_valid():
                insumo_seleccionado = form_insumo.cleaned_data['insumo']
                if not Formulacion.objects.filter(producto=producto, insumo=insumo_seleccionado).exists():
                    nuevo_item = form_insumo.save(commit=False)
                    nuevo_item.producto = producto
                    # Si se usa el margen de desperdicio predeterminado, asignarlo automáticamente
                    if usar_margen_desperdicio_predeterminado:
                        nuevo_item.porcentaje_desperdicio = mipyme_actual.margen_desperdicio_predeterminado
                    nuevo_item.save()
                # Redirigimos para limpiar el formulario y evitar reenvíos
                return redirect('produccion:detalle_producto', producto_id=producto.id)

        # --- LÓGICA PARA EL FORMULARIO DE PROCESOS ---
        elif 'submit_proceso' in request.POST:
            form_proceso = PasoDeProduccionForm(request.POST, mipyme=mipyme_actual)
            if form_proceso.is_valid():
                proceso_seleccionado = form_proceso.cleaned_data['proceso']
                if not PasoDeProduccion.objects.filter(producto=producto, proceso=proceso_seleccionado).exists():
                    nuevo_paso = form_proceso.save(commit=False)
                    nuevo_paso.producto = producto
                    nuevo_paso.save()
                # Redirigimos para limpiar el formulario y evitar reenvíos
                return redirect('produccion:detalle_producto', producto_id=producto.id)

    # Obtenemos los items para mostrarlos en las tablas
    formulacion_items = Formulacion.objects.filter(producto=producto).order_by('insumo__nombre')
    pasos_produccion = PasoDeProduccion.objects.filter(producto=producto).order_by('proceso__nombre')

    contexto = {
        'producto': producto,
        'formulacion_items': formulacion_items,
        'pasos_produccion': pasos_produccion,
        'form_insumo': form_insumo,  # Pasamos el formulario de insumos
        'form_proceso': form_proceso,  # Pasamos el formulario de procesos
        'usar_margen_desperdicio_predeterminado': usar_margen_desperdicio_predeterminado,
        'margen_desperdicio_predeterminado': mipyme_actual.margen_desperdicio_predeterminado,
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/detalle_producto.html', contexto)

@login_required
@mipyme_requerida
def lista_insumos(request):
    """
    Muestra una lista de todos los insumos de la Mipyme del usuario.
    """
    insumos = Insumo.objects.filter(mipyme=request.user.mipyme).order_by('nombre')
    contexto = {
        'insumos': insumos,
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/lista_insumos.html', contexto)


@login_required
@mipyme_requerida
def crear_insumo(request):
    """
    Gestiona la creación de un nuevo insumo.
    """
    if request.method == 'POST':
        form = InsumoForm(request.POST, mipyme=request.user.mipyme)
        if form.is_valid():
            nuevo_insumo = form.save(commit=False)
            nuevo_insumo.mipyme = request.user.mipyme # Asigna la Mipyme del usuario
            nuevo_insumo.save()
            return redirect('produccion:lista_insumos')
    else:
        form = InsumoForm(mipyme=request.user.mipyme)

    contexto = {
        'form': form,
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/crear_insumo.html', contexto)

@login_required
def editar_producto(request, producto_id):
    """
    Gestiona la edición de un producto existente.
    """
    # Seguridad: Obtenemos el producto asegurándonos que pertenece a la Mipyme del usuario
    producto = get_object_or_404(Producto, id=producto_id, mipyme=request.user.mipyme)

    mipyme = request.user.mipyme
    usar_porcentaje_predeterminado = mipyme.porcentaje_ganancia_predeterminado > 0

    if request.method == 'POST':
        # Pasamos 'instance=producto' para indicarle al formulario que estamos actualizando
        form = ProductoForm(request.POST, instance=producto, usar_porcentaje_predeterminado=usar_porcentaje_predeterminado, porcentaje_predeterminado=mipyme.porcentaje_ganancia_predeterminado)
        if form.is_valid():
            producto_editado = form.save(commit=False)
            # Si se usa porcentaje predeterminado, asignarlo automáticamente
            if usar_porcentaje_predeterminado:
                producto_editado.porcentaje_ganancia = mipyme.porcentaje_ganancia_predeterminado
            producto_editado.save()

            return redirect('produccion:lista_productos') # Redirigimos a la lista de productos
    else:
        # Si es GET, mostramos el formulario pre-llenado con los datos del producto
        form = ProductoForm(instance=producto, usar_porcentaje_predeterminado=usar_porcentaje_predeterminado, porcentaje_predeterminado=mipyme.porcentaje_ganancia_predeterminado)

    contexto = {
        'form': form,
        'producto': producto, # Lo pasamos para usarlo en el título de la plantilla
        'usar_porcentaje_predeterminado': usar_porcentaje_predeterminado,
        'porcentaje_predeterminado': mipyme.porcentaje_ganancia_predeterminado,
        'nombrepine': mipyme.nombre
    }
    # Reutilizaremos la plantilla de creación de productos
    return render(request, 'produccion/crear_producto.html', contexto)


@login_required
def eliminar_producto(request, producto_id):
    """
    Elimina un producto específico y su formulación asociada.
    """
    # Seguridad: Obtenemos el producto asegurándonos que pertenece a la Mipyme del usuario
    producto_a_borrar = get_object_or_404(Producto, id=producto_id, mipyme=request.user.mipyme)

    # Solo proceder si la solicitud es de tipo POST
    if request.method == 'POST':
        producto_a_borrar.delete()
        # Redirigir de vuelta a la lista de productos
        return redirect('produccion:lista_productos')

    # Si alguien intenta acceder a la URL con GET, simplemente lo redirigimos
    return redirect('produccion:lista_productos')


@login_required
def eliminar_formulacion_item(request, producto_id, item_id):
    """
    Elimina un insumo específico de la formulación de un producto.
    """
    # Verificación de seguridad: Asegurarnos de que el item a borrar
    # pertenece a un producto que a su vez pertenece a la Mipyme del usuario.
    item_a_borrar = get_object_or_404(
        Formulacion,
        id=item_id,
        producto__id=producto_id,
        producto__mipyme=request.user.mipyme
    )

    # Solo proceder si la solicitud es de tipo POST por seguridad
    if request.method == 'POST':
        item_a_borrar.delete()
        # Redirigir de vuelta a la página de detalle del producto
        return redirect('produccion:detalle_producto', producto_id=producto_id)

    # Si alguien intenta acceder a la URL con GET, simplemente lo redirigimos
    # sin hacer nada.
    return redirect('produccion:detalle_producto', producto_id=producto_id)

@login_required
def editar_formulacion_item(request, producto_id, item_id):
    """
    Edita la cantidad de un insumo en la formulación de un producto.
    """
    # Seguridad: Aseguramos que el item pertenece a un producto de la mipyme del usuario
    item = get_object_or_404(
        Formulacion,
        id=item_id,
        producto__id=producto_id,
        producto__mipyme=request.user.mipyme
    )

    if request.method == 'POST':
        # Pasamos 'instance=item' para que el formulario sepa qué objeto actualizar
        form = FormulacionUpdateForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            # Redirigimos de vuelta a la página de detalle del producto
            return redirect('produccion:detalle_producto', producto_id=producto_id)
    else:
        # Si es GET, creamos el formulario pre-llenado con los datos del item existente
        form = FormulacionUpdateForm(instance=item)

    contexto = {
        'form': form,
        'item': item,
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/editar_formulacion_item.html', contexto)


@login_required
def editar_insumo(request, insumo_id):
    """
    Gestiona la edición de un insumo existente.
    """
    # Seguridad: Obtenemos el insumo asegurándonos que pertenece a la Mipyme del usuario
    insumo = get_object_or_404(Insumo, id=insumo_id, mipyme=request.user.mipyme)

    if request.method == 'POST':
        # Pasamos 'instance=insumo' para indicarle al formulario que estamos actualizando
        form = InsumoForm(request.POST, instance=insumo, mipyme=request.user.mipyme)
        if form.is_valid():
            form.save()
            return redirect('produccion:lista_insumos') # Redirigimos a la lista
    else:
        # Si es GET, mostramos el formulario pre-llenado con los datos del insumo
        form = InsumoForm(instance=insumo, mipyme=request.user.mipyme)

    contexto = {
        'form': form,
        'insumo': insumo, # Pasamos el insumo para poder usar su nombre en el título
        'nombrepine': request.user.mipyme.nombre
    }
    # Reutilizaremos la plantilla de creación, ya que el formulario es el mismo
    return render(request, 'produccion/crear_insumo.html', contexto)


@login_required
def eliminar_insumo(request, insumo_id):
    """
    Elimina un insumo específico.
    """
    # Seguridad: Obtenemos el insumo asegurándonos que pertenece a la Mipyme del usuario
    insumo_a_borrar = get_object_or_404(Insumo, id=insumo_id, mipyme=request.user.mipyme)

    # Solo proceder si la solicitud es de tipo POST
    if request.method == 'POST':
        insumo_a_borrar.delete()
        # Redirigir de vuelta a la lista de insumos
        return redirect('produccion:lista_insumos')

    # Si alguien intenta acceder a la URL con GET, simplemente lo redirigimos
    return redirect('produccion:lista_insumos')


# --- INICIO DE NUEVAS VISTAS PARA PROCESOS ---

@login_required
@mipyme_requerida
def lista_procesos(request):
    """
    Muestra una lista de todos los procesos de la Mipyme del usuario.
    """
    procesos = Proceso.objects.filter(mipyme=request.user.mipyme).order_by('nombre')
    contexto = {
        'procesos': procesos,
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/lista_procesos.html', contexto)

@login_required
@mipyme_requerida
@rol_requerido('ADMIN', 'EDITOR')
def crear_proceso(request):
    """
    Gestiona la creación de un nuevo proceso.
    """
    if request.method == 'POST':
        form = ProcesoForm(request.POST)
        if form.is_valid():
            nuevo_proceso = form.save(commit=False)
            nuevo_proceso.mipyme = request.user.mipyme
            nuevo_proceso.save()
            return redirect('produccion:lista_procesos')
    else:
        form = ProcesoForm()

    contexto = {'form': form}
    return render(request, 'produccion/crear_proceso.html', contexto)


@login_required
@rol_requerido('ADMIN', 'EDITOR')
def editar_proceso(request, proceso_id):
    """
    Gestiona la edición de un proceso existente.
    """
    proceso = get_object_or_404(Proceso, id=proceso_id, mipyme=request.user.mipyme)
    if request.method == 'POST':
        form = ProcesoForm(request.POST, instance=proceso)
        if form.is_valid():
            form.save()
            return redirect('produccion:lista_procesos')
    else:
        form = ProcesoForm(instance=proceso)
    contexto = {'form': form, 'proceso': proceso}
    return render(request, 'produccion/crear_proceso.html', contexto)


@login_required
@rol_requerido('ADMIN')
def eliminar_proceso(request, proceso_id):
    """
    Elimina un proceso específico.
    """
    proceso_a_borrar = get_object_or_404(Proceso, id=proceso_id, mipyme=request.user.mipyme)
    if request.method == 'POST':
        proceso_a_borrar.delete()
        return redirect('produccion:lista_procesos')
    return redirect('produccion:lista_procesos')

# --- FIN DE NUEVAS VISTAS PARA PROCESOS ---


# --- INICIO DE NUEVAS VISTAS PARA PASOS DE PRODUCCIÓN ---

@login_required
def editar_paso_produccion(request, producto_id, paso_id):
    """
    Edita el tiempo de un proceso en la ruta de producción de un producto.
    """
    paso = get_object_or_404(
        PasoDeProduccion, id=paso_id, producto__id=producto_id, producto__mipyme=request.user.mipyme
    )
    if request.method == 'POST':
        form = PasoUpdateForm(request.POST, instance=paso)
        if form.is_valid():
            form.save()
            return redirect('produccion:detalle_producto', producto_id=producto_id)
    else:
        form = PasoUpdateForm(instance=paso)

    contexto = {'form': form,
                'paso': paso,
                'nombrepine': request.user.mipyme.nombre
                }
    return render(request, 'produccion/editar_paso_produccion.html', contexto)


@login_required
def eliminar_paso_produccion(request, producto_id, paso_id):
    """
    Elimina un proceso de la ruta de producción de un producto.
    """
    paso_a_borrar = get_object_or_404(
        PasoDeProduccion, id=paso_id, producto__id=producto_id, producto__mipyme=request.user.mipyme
    )
    if request.method == 'POST':
        paso_a_borrar.delete()
        return redirect('produccion:detalle_producto', producto_id=producto_id)
    return redirect('produccion:detalle_producto', producto_id=producto_id)

# --- FIN DE NUEVAS VISTAS PARA PASOS DE PRODUCCIÓN ---


@login_required
@mipyme_requerida
def calculadora_lotes(request, producto_id):
    """
    Calculadora para estimar cantidades de insumos y costos para un lote de producción.
    Permite producir el lote restando insumos del inventario.
    """
    producto = get_object_or_404(Producto, id=producto_id, mipyme=request.user.mipyme)

    resultados = None
    costo_total_insumos = 0
    costo_total_procesos = 0
    ingresos_estimados = 0
    ganancia_estimada = 0
    error_stock = None
    produccion_exitosa = False

    if request.method == 'POST':
        form = CalculadoraLotesForm(request.POST)
        if form.is_valid():
            cantidad_unidades = form.cleaned_data['cantidad_unidades']

            # Calcular insumos para el lote
            resultados = []
            insumos_a_restar = []
            for item in producto.formulacion.all():
                cantidad_total = item.cantidad * cantidad_unidades
                cantidad_con_desperdicio = cantidad_total * (1 + (item.porcentaje_desperdicio / 100))
                costo_unitario = item.insumo.costo_unitario
                costo_con_desperdicio = cantidad_total * costo_unitario * (1 + (item.porcentaje_desperdicio / 100))
                costo_total_insumos += costo_con_desperdicio
                resultados.append({
                    'insumo': item.insumo.nombre,
                    'unidad': item.insumo.unidad.abreviatura,
                    'cantidad_por_unidad': item.cantidad,
                    'cantidad_total': cantidad_total,
                    'costo': costo_con_desperdicio,
                })
                insumos_a_restar.append({
                    'insumo': item.insumo,
                    'cantidad_a_restar': cantidad_con_desperdicio,
                })

            # Calcular procesos para el lote
            for paso in producto.pasodeproduccion_set.all():
                costo_proceso_lote = (decimal.Decimal(paso.tiempo_en_minutos) / 60) * paso.proceso.costo_por_hora * cantidad_unidades
                costo_total_procesos += costo_proceso_lote

            # Calcular ingresos y ganancia
            ingresos_estimados = producto.precio_venta * cantidad_unidades
            ganancia_estimada = producto.margen_de_ganancia * cantidad_unidades

            # Si se solicita producir el lote
            if 'producir_lote' in request.POST:
                # Validar stock de insumos
                for insumo_data in insumos_a_restar:
                    if insumo_data['insumo'].stock_actual < insumo_data['cantidad_a_restar']:
                        error_stock = f"No hay suficiente stock de {insumo_data['insumo'].nombre}. Disponible: {insumo_data['insumo'].stock_actual} {insumo_data['insumo'].unidad.abreviatura}, requerido: {insumo_data['cantidad_a_restar']:.2f} {insumo_data['insumo'].unidad.abreviatura}."
                        break
                else:
                    # Restar stock de insumos
                    for insumo_data in insumos_a_restar:
                        insumo_data['insumo'].stock_actual -= insumo_data['cantidad_a_restar']
                        insumo_data['insumo'].save(update_fields=['stock_actual'])

                    # Aumentar stock del producto
                    producto.stock_actual += cantidad_unidades
                    producto.save(update_fields=['stock_actual'])

                    produccion_exitosa = True
                    # Redirigir con mensaje de éxito
                    from django.contrib import messages
                    messages.success(request, f'Se produjo exitosamente {cantidad_unidades} unidades de {producto.nombre}.')
                    return redirect('produccion:detalle_producto', producto_id=producto.id)
    else:
        form = CalculadoraLotesForm()

    costo_total = costo_total_insumos + costo_total_procesos

    contexto = {
        'producto': producto,
        'form': form,
        'resultados': resultados,
        'costo_total_insumos': costo_total_insumos,
        'costo_total_procesos': costo_total_procesos,
        'costo_total': costo_total,
        'ingresos_estimados': ingresos_estimados,
        'ganancia_estimada': ganancia_estimada,
        'error_stock': error_stock,
        'produccion_exitosa': produccion_exitosa,
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/calculadora_lotes.html', contexto)

# --- FIN DE CALCULADORA ---


# --- INICIO DE VENTAS (FACTURACIÓN) ---

@login_required
@mipyme_requerida
def registrar_venta(request):
    """
    Registrar una nueva venta con múltiples productos usando un ModelFormSet de VentaItem.
    Realiza:
     - Validación de stock
     - Resta automática del stock
     - Cálculo de subtotales y total
     - Devuelve datos para modal de confirmación e impresión
    """
    # Datos de productos para JS
    productos = Producto.objects.filter(mipyme=request.user.mipyme).values('id', 'nombre', 'precio_venta')
    productos_data = {
        str(p['id']): {'nombre': p['nombre'], 'precio_venta': float(p['precio_venta'])}
        for p in productos
    }
    productos_json = json.dumps(productos_data)

    if request.method == 'POST':
        formset = VentaItemFormSet(
            request.POST,
            form_kwargs={'mipyme': request.user.mipyme},
            queryset=VentaItem.objects.none()
        )
        if formset.is_valid():
            venta = Venta.objects.create(mipyme=request.user.mipyme)
            items_data = []
            total_items = 0

            for form in formset:
                if not form.cleaned_data or form.cleaned_data.get('DELETE'):
                    continue

                producto = form.cleaned_data['producto']
                cantidad = form.cleaned_data['cantidad']

                # Validación de stock por seguridad (además del clean del form)
                if cantidad > producto.stock_actual:
                    venta.delete()
                    contexto = {
                        'formset': formset,
                        'titulo': 'Registrar Venta',
                        'nombrepine': request.user.mipyme.nombre,
                        'productos_json': productos_json,
                        'error': f"No hay suficiente stock para {producto.nombre}. Stock disponible: {producto.stock_actual} unidades."
                    }
                    return render(request, 'produccion/registrar_venta.html', contexto)

                # Crear item con precio actual del producto
                item = VentaItem(
                    venta=venta,
                    producto=producto,
                    cantidad=cantidad,
                    precio_unitario=producto.precio_venta
                )
                item.save()

                # Restar del stock del producto
                producto.stock_actual -= cantidad
                producto.save(update_fields=['stock_actual'])

                items_data.append({
                    'producto': producto.nombre,
                    'cantidad': cantidad,
                    'precio_unitario': float(item.precio_unitario),
                    'subtotal': float(item.subtotal),
                })
                total_items += 1

            if total_items == 0:
                venta.delete()
                contexto = {
                    'formset': formset,
                    'titulo': 'Registrar Venta',
                    'nombrepine': request.user.mipyme.nombre,
                    'productos_json': productos_json,
                    'error': 'Debe agregar al menos un producto a la venta.'
                }
                return render(request, 'produccion/registrar_venta.html', contexto)

            # Calcular total de la venta
            venta.calcular_total()

            venta_dict = {
                'id': venta.id,
                'fecha': venta.fecha.strftime('%d/%m/%Y %H:%M'),
                'total': float(venta.total),
                'items': items_data
            }
            venta_json = json.dumps(venta_dict)

            # Render con formulario limpio, datos para modal y JSON embebido
            contexto = {
                'formset': VentaItemFormSet(queryset=VentaItem.objects.none(), form_kwargs={'mipyme': request.user.mipyme}),
                'titulo': 'Registrar Venta',
                'nombrepine': request.user.mipyme.nombre,
                'productos_json': productos_json,
                'venta_exitosa': True,
                'venta_json': venta_json,
            }
            return render(request, 'produccion/registrar_venta.html', contexto)
        else:
            contexto = {
                'formset': formset,
                'titulo': 'Registrar Venta',
                'nombrepine': request.user.mipyme.nombre,
                'productos_json': productos_json,
            }
            return render(request, 'produccion/registrar_venta.html', contexto)

    # GET
    formset = VentaItemFormSet(queryset=VentaItem.objects.none(), form_kwargs={'mipyme': request.user.mipyme})
    contexto = {
        'formset': formset,
        'titulo': 'Registrar Venta',
        'nombrepine': request.user.mipyme.nombre,
        'productos_json': productos_json,
    }
    return render(request, 'produccion/registrar_venta.html', contexto)


@login_required
@mipyme_requerida
def historial_ventas(request):
    """
    Lista el historial de ventas de la MIPYME.
    """
    ventas = Venta.objects.filter(mipyme=request.user.mipyme).prefetch_related('items__producto').order_by('-fecha')
    contexto = {
        'ventas': ventas,
        'titulo': 'Historial de Ventas',
        'nombrepine': request.user.mipyme.nombre,
    }
    return render(request, 'produccion/historial_ventas.html', contexto)

# --- CONFIGURACIÓN ---

@login_required
@mipyme_requerida
def configuracion(request):
    """
    Vista para la configuración de la aplicación de producción.
    """
    contexto = {
        'titulo': 'Configuración',
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/configuracion.html', contexto)


@login_required
@mipyme_requerida
def cambiar_contrasena(request):
    """
    Vista para cambiar la contraseña del usuario.
    """
    if request.method == 'POST':
        form = CambiarContrasenaForm(request.POST)
        if form.is_valid():
            # Verificar contraseña actual
            if not request.user.check_password(form.cleaned_data['password_actual']):
                messages.error(request, 'La contraseña actual es incorrecta.')
            else:
                # Cambiar contraseña
                request.user.set_password(form.cleaned_data['password_nueva'])
                request.user.save()
                # Mantener la sesión activa
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Contraseña cambiada exitosamente.')
                return redirect('produccion:configuracion')
    else:
        form = CambiarContrasenaForm()

    contexto = {
        'form': form,
        'titulo': 'Cambiar Contraseña',
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/configuraciones/cambiar_contrasena.html', contexto)


@login_required
@mipyme_requerida
def actualizar_perfil(request):
    """
    Vista para actualizar la información de contacto del usuario.
    """
    if request.method == 'POST':
        form = ActualizarPerfilForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
            return redirect('produccion:configuracion')
    else:
        form = ActualizarPerfilForm(instance=request.user)

    contexto = {
        'form': form,
        'titulo': 'Actualizar Perfil',
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/configuraciones/actualizar_perfil.html', contexto)


@login_required
@mipyme_requerida
def configurar_avatar(request):
    """
    Vista para configurar el avatar del usuario.
    """
    if request.method == 'POST':
        form = ConfigurarAvatarForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Avatar actualizado exitosamente.')
            return redirect('produccion:configuracion')
    else:
        form = ConfigurarAvatarForm(instance=request.user)

    contexto = {
        'form': form,
        'titulo': 'Configurar Avatar',
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/configuraciones/configurar_avatar.html', contexto)


@login_required
@mipyme_requerida
@rol_requerido('ADMIN')
def editar_informacion_empresa(request):
    """
    Vista para editar la información básica de la empresa (MiPyme).
    Solo accesible para administradores.
    """
    mipyme = request.user.mipyme

    if request.method == 'POST':
        form = EditarInformacionEmpresaForm(request.POST, instance=mipyme)
        if form.is_valid():
            form.save()
            messages.success(request, 'Información de la empresa actualizada exitosamente.')
            return redirect('produccion:configuracion')
    else:
        form = EditarInformacionEmpresaForm(instance=mipyme)

    contexto = {
        'form': form,
        'titulo': 'Editar Información de la Empresa',
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/configuraciones/editar_informacion_empresa.html', contexto)


@login_required
@mipyme_requerida
@rol_requerido('ADMIN')
def configurar_logo(request):
    """
    Vista para configurar el logo de la empresa.
    Solo accesible para administradores.
    """
    mipyme = request.user.mipyme

    if request.method == 'POST':
        form = ConfigurarLogoForm(request.POST, request.FILES, instance=mipyme)
        if form.is_valid():
            form.save()
            messages.success(request, 'Logo de la empresa actualizado exitosamente.')
            return redirect('produccion:configuracion')
    else:
        form = ConfigurarLogoForm(instance=mipyme)

    contexto = {
        'form': form,
        'titulo': 'Configurar Logo de la Empresa',
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/configuraciones/configurar_logo.html', contexto)


@login_required
@mipyme_requerida
@rol_requerido('ADMIN')
def cambiar_sector_economico(request):
    """
    Vista para cambiar el sector económico de la empresa.
    Solo accesible para administradores.
    """
    mipyme = request.user.mipyme

    if request.method == 'POST':
        form = CambiarSectorEconomicoForm(request.POST, instance=mipyme)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sector económico actualizado exitosamente.')
            return redirect('produccion:configuracion')
    else:
        form = CambiarSectorEconomicoForm(instance=mipyme)

    contexto = {
        'form': form,
        'titulo': 'Cambiar Sector Económico',
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/configuraciones/cambiar_sector_economico.html', contexto)


@login_required
@mipyme_requerida
@rol_requerido('ADMIN')
def configurar_parametros_produccion(request):
    """
    Vista para configurar los parámetros de producción de la empresa.
    Solo accesible para administradores.
    """
    mipyme = request.user.mipyme

    if request.method == 'POST':
        form = ConfigurarParametrosProduccionForm(request.POST, instance=mipyme)
        if form.is_valid():
            form.save()
            messages.success(request, 'Parámetros de producción actualizados exitosamente.')
            return redirect('produccion:configuracion')
    else:
        form = ConfigurarParametrosProduccionForm(instance=mipyme)

    contexto = {
        'form': form,
        'titulo': 'Configurar Parámetros de Producción',
        'nombrepine': request.user.mipyme.nombre
    }
    return render(request, 'produccion/configuraciones/configurar_parametros_produccion.html', contexto)

@login_required
@mipyme_requerida
def exportar_productos_excel(request):
    """
    Exporta todos los productos de la Mipyme del usuario a un archivo Excel,
    incluyendo datos relevantes y relacionados.
    """
    mipyme = request.user.mipyme
    productos = Producto.objects.filter(mipyme=mipyme).prefetch_related('formulacion__insumo', 'pasodeproduccion_set__proceso', 'estándares').order_by('nombre')

    # Crear workbook
    wb = Workbook()
    ws_productos = wb.active
    ws_productos.title = "Productos"

    # Hoja de Productos
    headers_productos = [
        'Nombre', 'Descripción', 'Precio Venta', 'Porcentaje Ganancia', 'Stock Actual',
        'Costo Producción', 'Costo Insumos', 'Costo Procesos', 'Margen Ganancia',
        'Peso (kg)', 'Largo (cm)', 'Ancho (cm)', 'Alto (cm)', 'Presentación'
    ]
    ws_productos.append(headers_productos)

    for producto in productos:
        row = [
            producto.nombre,
            producto.descripcion or '',
            float(producto.precio_venta) if producto.precio_venta else 0,
            float(producto.porcentaje_ganancia) if producto.porcentaje_ganancia else 0,
            producto.stock_actual,
            float(producto.costo_de_produccion),
            float(producto.costo_insumos),
            float(producto.costo_procesos),
            float(producto.margen_de_ganancia),
            float(producto.peso) if producto.peso else '',
            float(producto.tamano_largo) if producto.tamano_largo else '',
            float(producto.tamano_ancho) if producto.tamano_ancho else '',
            float(producto.tamano_alto) if producto.tamano_alto else '',
            producto.presentacion or ''
        ]
        ws_productos.append(row)

    # Hoja de Formulación (Insumos)
    ws_formulacion = wb.create_sheet("Formulación")
    headers_formulacion = ['Producto', 'Insumo', 'Cantidad', 'Unidad', 'Costo Unitario', 'Porcentaje Desperdicio']
    ws_formulacion.append(headers_formulacion)

    for producto in productos:
        for item in producto.formulacion.all():
            row = [
                producto.nombre,
                item.insumo.nombre,
                float(item.cantidad),
                item.insumo.unidad.abreviatura,
                float(item.insumo.costo_unitario),
                float(item.porcentaje_desperdicio)
            ]
            ws_formulacion.append(row)

    # Hoja de Procesos
    ws_procesos = wb.create_sheet("Procesos")
    headers_procesos = ['Producto', 'Proceso', 'Tiempo (minutos)', 'Costo por Hora']
    ws_procesos.append(headers_procesos)

    for producto in productos:
        for paso in producto.pasodeproduccion_set.all():
            row = [
                producto.nombre,
                paso.proceso.nombre,
                paso.tiempo_en_minutos,
                float(paso.proceso.costo_por_hora)
            ]
            ws_procesos.append(row)

    # Hoja de Estándares (si existen)
    ws_estandares = wb.create_sheet("Estándares")
    headers_estandares = [
        'Producto', 'Peso Mín', 'Peso Máx', 'Largo Mín', 'Largo Máx',
        'Ancho Mín', 'Ancho Máx', 'Alto Mín', 'Alto Máx', 'Presentación Estándar'
    ]
    ws_estandares.append(headers_estandares)

    for producto in productos:
        if hasattr(producto, 'estándares') and producto.estándares:
            est = producto.estándares
            row = [
                producto.nombre,
                float(est.peso_min) if est.peso_min else '',
                float(est.peso_max) if est.peso_max else '',
                float(est.tamano_largo_min) if est.tamano_largo_min else '',
                float(est.tamano_largo_max) if est.tamano_largo_max else '',
                float(est.tamano_ancho_min) if est.tamano_ancho_min else '',
                float(est.tamano_ancho_max) if est.tamano_ancho_max else '',
                float(est.tamano_alto_min) if est.tamano_alto_min else '',
                float(est.tamano_alto_max) if est.tamano_alto_max else '',
                est.presentacion_estandar or ''
            ]
            ws_estandares.append(row)

    # Preparar respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="productos_{mipyme.nombre.replace(" ", "_")}.xlsx"'

    # Guardar workbook en la respuesta
    wb.save(response)
    return response

# --- FIN DE CONFIGURACIÓN ---
# --- FIN DE VENTAS ---